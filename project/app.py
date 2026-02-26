from flask import Flask, render_template, request, redirect, url_for, flash
import sqlite3
import os
import csv
import chardet
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from flask import send_file
app = Flask(__name__)
app.secret_key = "secret123"

DB_NAME = "database.db"
UPLOAD_FOLDER = "uploads"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ================= DATABASE =================
def get_db():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn


def create_table():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS students(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                age INTEGER NOT NULL
            )
        """)

# ================= HOME =================
@app.route("/")
def index():
    search = request.args.get("search", "")

    conn = get_db()

    if search:
        students = conn.execute(
            "SELECT * FROM students WHERE name LIKE ?",
            ('%' + search + '%',)
        ).fetchall()
    else:
        students = conn.execute("SELECT * FROM students").fetchall()

    count = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    conn.close()

    return render_template("index.html",
                           students=students,
                           count=count,
                           search=search)

# ================= ADD =================
@app.route("/add", methods=["GET", "POST"])
def add_student():
    if request.method == "POST":
        name = request.form["name"]
        age = request.form["age"]

        with get_db() as conn:
            conn.execute(
                "INSERT INTO students(name, age) VALUES (?, ?)",
                (name, age)
            )

        flash("Student Added Successfully ✅")
        return redirect(url_for("index"))

    return render_template("add.html")

# ================= EDIT =================
@app.route("/edit/<int:id>", methods=["GET", "POST"])
def edit_student(id):

    conn = get_db()

    if request.method == "POST":
        name = request.form["name"]
        age = request.form["age"]

        conn.execute(
            "UPDATE students SET name=?, age=? WHERE id=?",
            (name, age, id)
        )
        conn.commit()
        conn.close()

        flash("Student Updated ✏️")
        return redirect(url_for("index"))

    student = conn.execute(
        "SELECT * FROM students WHERE id=?", (id,)
    ).fetchone()

    conn.close()
    return render_template("edit.html", student=student)

# ================= DELETE =================
@app.route("/delete/<int:id>")
def delete_student(id):
    with get_db() as conn:
        conn.execute("DELETE FROM students WHERE id=?", (id,))

    flash("Student Deleted 🗑️")
    return redirect(url_for("index"))

# ================= UNIVERSAL FILE UPLOAD =================
@app.route("/upload_universal", methods=["POST"])
def upload_universal():

    file = request.files.get("file")

    if not file or file.filename == "":
        flash("No file selected ❌")
        return redirect(url_for("index"))

    filename = file.filename.lower()
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    inserted = 0
    conn = get_db()

    try:

        # ---------- EXCEL ----------
        if filename.endswith(".xlsx"):
            wb = load_workbook(filepath)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue

                conn.execute(
                    "INSERT INTO students(name, age) VALUES (?, ?)",
                    (str(row[0]), int(row[1]))
                )
                inserted += 1

        # ---------- XML ----------
        elif filename.endswith(".xml"):
            tree = ET.parse(filepath)
            root = tree.getroot()

            for student in root.findall("student"):
                name = student.find("name").text
                age = student.find("age").text

                conn.execute(
                    "INSERT INTO students(name, age) VALUES (?, ?)",
                    (name, age)
                )
                inserted += 1

        # ---------- CSV ----------
        elif filename.endswith(".csv"):

            with open(filepath, "rb") as f:
                encoding = chardet.detect(f.read())["encoding"]

            with open(filepath, "r", encoding=encoding) as f:
                reader = csv.reader(f)
                next(reader, None)

                for row in reader:
                    conn.execute(
                        "INSERT INTO students(name, age) VALUES (?, ?)",
                        (row[0], row[1])
                    )
                    inserted += 1

        # ---------- TXT ----------
        elif filename.endswith(".txt"):

            with open(filepath, "rb") as f:
                encoding = chardet.detect(f.read())["encoding"]

            with open(filepath, "r", encoding=encoding) as f:
                for line in f:
                    if not line.strip():
                        continue

                    name, age = line.strip().split(",")

                    conn.execute(
                        "INSERT INTO students(name, age) VALUES (?, ?)",
                        (name, age)
                    )
                    inserted += 1

        else:
            flash("Unsupported file type ❌")
            conn.close()
            return redirect(url_for("index"))

        conn.commit()
        conn.close()

        flash(f"✅ Import Successful — {inserted} records added")

    except Exception as e:
        conn.close()
        flash(f"Import Error: {str(e)}")

    return redirect(url_for("index"))
# ================= DOWNLOAD EXCEL =================
@app.route("/download_excel")
def download_excel():

    conn = get_db()
    students = conn.execute("SELECT * FROM students").fetchall()
    conn.close()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    # Header row
    ws.append(["ID", "Name", "Age"])

    # Data rows
    for s in students:
        ws.append([s["id"], s["name"], s["age"]])

    filepath = os.path.join(UPLOAD_FOLDER, "students_export.xlsx")
    wb.save(filepath)

    return send_file(
        filepath,
        as_attachment=True,
        download_name="students.xlsx"
    )
# ================= MAIN =================
if __name__ == "__main__":
    create_table()
    app.run(host="0.0.0.0", port=5000)